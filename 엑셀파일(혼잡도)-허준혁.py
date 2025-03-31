# 엑셀파일 과제
# 파일 불러오기
from google.colab import files

f = files.upload()

# 엑셀 파일을 다루는 라이브러리 선언
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

import csv

with open('data1.csv', encoding= 'cp949') as f:
  reader = csv.reader(f)
  for row in reader:
    ws.append(row)

ws2 = wb.create_sheet('Sheet2')

with open('data2.csv', encoding= 'cp949') as f:
  reader = csv.reader(f)
  for row in reader:
    ws2.append(row)

ws3 = wb.create_sheet('Sheet3')

with open('data3.csv', encoding= 'cp949') as f:
  reader = csv.reader(f)
  for row in reader:
    ws3.append(row)

ws.title = '경기도 의정부시_의정부경전철 혼잡도'
ws2.title = '대전교통공사_열차 혼잡도 분석'
ws3.title = '서울 특별시_지하철 혼잡도 정보'

wb.save('traffic.xlsx')
